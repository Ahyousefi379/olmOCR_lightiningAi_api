import os
from pathlib import Path
from typing import List, Dict
import time

from openpyxl import Workbook, load_workbook

from pdf_references_cleaner import remove_pages_after_references
from merger_ocrs_util import symmetrical_merge
from validate_pdf import validate_pdfs
from headings_fix_util import ScientificTextToMarkdownConverter
from olmOCR_api_util import OlmOcrPollingClient



# --- CONFIGURATION ---
BASE_DIR = Path("H://python_projects//scientific//olmOCR//pdfs")
RAW_PDFS_DIR = BASE_DIR / "raw"
CLEANED_PDFS_DIR = BASE_DIR / "cleaned"
RAW_OCR_DIR = BASE_DIR / "raw ocr"
MD_DIR = BASE_DIR / "md"

SERVER_URL = "https://8000-01k4ymt1036tngva09ywdk72z6.cloudspaces.litng.ai"
POLL_INTERVAL_SECONDS = 10
TOTAL_WAIT_MINUTES = 15
NUM_OCR_VERSIONS = 1
MAX_RETRY_ATTEMPTS = 2


def setup_directories():
    print("Setting up directories...")
    for dir_path in [CLEANED_PDFS_DIR, RAW_OCR_DIR, MD_DIR]:
        dir_path.mkdir(parents=True, exist_ok=True)


def clean_pdfs(source_dir: Path, target_dir: Path):
    print(f"Scanning for PDFs in '{source_dir}'...")
    raw_pdf_files = list(source_dir.glob('*.pdf'))

    if not raw_pdf_files:
        print(f"No PDF files found in '{source_dir}'.")
        return

    print(f"Found {len(raw_pdf_files)} PDFs. Cleaning references...")
    for pdf_path in raw_pdf_files:
        output_path = target_dir / f"_cleaned_{pdf_path.name}"
        print(f"  - Processing {pdf_path.name}")
        remove_pages_after_references(
            input_pdf_path=str(pdf_path),
            output_pdf_path=str(output_path)
        )
    print("PDF cleaning complete.\n")



def is_ocr_successful(output_path: Path) -> bool:
    if not output_path.exists():
        return False
    try:
        with open(output_path, 'r', encoding='utf-8') as f:
            content = f.read().strip()
            return len(content) > 100
    except Exception:
        return False


def get_base_name_from_cleaned(pdf_path: Path) -> str:
    return pdf_path.stem.replace('_cleaned_', '')


def perform_ocr_with_validation(source_files: List[Path], target_dir: Path, save_reports_dir:Path):
    print("Starting OCR process...")
    if not source_files:
        print("No valid cleaned PDFs to process.")
        return

    client = OlmOcrPollingClient(
        base_url=SERVER_URL,
        json_report_file=str(save_reports_dir / "ocr_report.json")
    )

    successful_pdfs = []

    # Excel path to record failed PDFs
    excel_path = save_reports_dir / "failed_ocr_pdfs.xlsx"
    
    # Initialize Excel if it doesn't exist
    if not excel_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Failed OCR PDFs"
        ws.append(["Filename", "Reason"])
        wb.save(str(excel_path))

    for pdf_path in source_files:
        base_name = get_base_name_from_cleaned(pdf_path)
        cleaned_filename = pdf_path.name

        print(f"  - Processing '{cleaned_filename}' (need {NUM_OCR_VERSIONS} successful versions)...")

        existing_successes = client.get_existing_success_count(cleaned_filename)
        print(f"    - Found {existing_successes} existing successful attempts")

        successful_versions = existing_successes
        failed_reason = ""

        for version in range(existing_successes + 1, NUM_OCR_VERSIONS + 1):
            output_path = target_dir / f"{base_name}_attempt_{version}.md"

            print(f"    - Working on version {version}/{NUM_OCR_VERSIONS}")

            version_successful = False
            for attempt in range(1, MAX_RETRY_ATTEMPTS + 1):
                print(f"      - Retry {attempt}: Processing version {version}")

                try:
                    success = client.process_document(
                        file_path=str(pdf_path),
                        save_output=True,
                        output_dir=str(output_path.parent),
                        output_filename=output_path.name,
                        poll_interval_sec=POLL_INTERVAL_SECONDS,
                        total_wait_min=TOTAL_WAIT_MINUTES,
                        attempt_id=f"attempt_{version}",
                        force_process=True
                    )

                    if success and is_ocr_successful(output_path):
                        print(f"      - Version {version} completed successfully")
                        version_successful = True
                        successful_versions += 1
                        break
                    else:
                        failed_reason = "No valid OCR output"
                        print(f"      - Version {version} failed (no valid output)")

                except Exception as e:
                    failed_reason = str(e)
                    print(f"      - Version {version} failed with error: {failed_reason}")

                if attempt < MAX_RETRY_ATTEMPTS:
                    print(f"      - Retrying version {version} in 10 seconds...")
                    time.sleep(10)

            if not version_successful:
                print(f"    - Version {version} failed after {MAX_RETRY_ATTEMPTS} attempts")

        if successful_versions >= NUM_OCR_VERSIONS:
            print(f"  ✓ '{cleaned_filename}' completed successfully ({successful_versions}/{NUM_OCR_VERSIONS} versions)")
            successful_pdfs.append(cleaned_filename)
        else:
            print(f"  ✗ '{cleaned_filename}' failed - only {successful_versions}/{NUM_OCR_VERSIONS} versions succeeded")
            
            # --- Record failed PDF in Excel immediately ---
            try:
                wb = load_workbook(str(excel_path))
                ws = wb.active
                ws.append([cleaned_filename, failed_reason])
                wb.save(str(excel_path))
                print(f"    - Recorded failed PDF in Excel: {excel_path.name}")
            except Exception as e:
                print(f"    - Could not write failed PDF to Excel: {e}")

    print(f"\nOCR processing complete!")
    print(f"Successful PDFs: {len(successful_pdfs)}")
    
    # Optionally, print a summary of failed PDFs from Excel
    if excel_path.exists():
        print(f"Failed PDFs recorded in Excel: {excel_path}")
    print()


def post_process_ocr(source_dir: Path, target_dir: Path):
    print("Starting OCR post-processing...")
    attempt_files: Dict[str, List[Path]] = {}
    for md_file in source_dir.glob("*_attempt*.md"):
        parts = md_file.stem.split("_attempt")
        if len(parts) == 2:
            base_name = parts[0]
            attempt_files.setdefault(base_name, []).append(md_file)

    if not attempt_files:
        print(f"No OCR attempt markdown files found in '{source_dir}'.")
        return

    for base_name, md_files in attempt_files.items():
        print(f"  - Processing '{base_name}'...")

        if len(md_files) < NUM_OCR_VERSIONS:
            print(f"    - Warning: Expected {NUM_OCR_VERSIONS} attempts, found {len(md_files)}. Using available ones.")

        md_files.sort(key=lambda x: int((str(x).split("attempt")[-1])[1]))

        texts = []
        for md_file in md_files[:NUM_OCR_VERSIONS]:
            try:
                with open(md_file, "r", encoding="utf-8") as f:
                    content = f.read()
                    if content.strip():
                        texts.append(content)
            except Exception as e:
                print(f"    - Warning: Could not read {md_file}: {e}")

        if not texts:
            print(f"    - Skipping '{base_name}', no valid text found.")
            continue

        merged_text = texts[0] if len(texts) == 1 else symmetrical_merge(text1=texts[0], text2=texts[1])

        output_path = target_dir / f"_final_{base_name}.md"
        print(f"    - Fixing headings and saving into '{output_path.name}'")

        try:
            convertor = ScientificTextToMarkdownConverter(merged_text)
            fixed_heading_text = convertor.convert()
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(fixed_heading_text)
        except Exception as e:
            print(f"    - Error processing '{base_name}': {e}")

    print("Post-processing complete.\n")


def main():
    setup_directories()

    # Step 1: Remove reference pages
    clean_pdfs(source_dir=RAW_PDFS_DIR,
               target_dir=CLEANED_PDFS_DIR)

    # Step 2: Validate PDFs
    valid_pdfs = validate_pdfs(source_dir=CLEANED_PDFS_DIR,
                               save_report_dir=BASE_DIR)

    # Step 3: Run OCR only on valid PDFs
    perform_ocr_with_validation(source_files=valid_pdfs,
                                target_dir=RAW_OCR_DIR,
                                save_reports_dir=BASE_DIR)

    # Step 4: Merge OCR results
    #post_process_ocr(source_dir=RAW_OCR_DIR, target_dir=MD_DIR)

    print("Pipeline finished.")


if __name__ == "__main__":
    main()
